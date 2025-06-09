import datetime
import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from core.constants import CYRILLIC_TO_LATIN, DEFAULT_INDEX, INDEX_MAP


class ExcelReader:
    def get_details(self, file_path: Path) -> tuple[list[str], str]:
        workbook = load_workbook(file_path, read_only=True, keep_links=False)
        sheet_names = workbook.sheetnames
        operator = workbook.properties.lastModifiedBy
        workbook.close()
        return sheet_names, operator

    def read_sheet(self, file_path: Path, sheet_name: str) -> pd.DataFrame:
        return (
            pd.read_excel(**self._get_kwargs(file_path, sheet_name))
            .dropna(axis=0)
        )

    @staticmethod
    def _get_kwargs(filepath: Path, sheet: str) -> dict:
        base = {'io': filepath, 'sheet_name': sheet}
        special = {
            'declaration_form': {
                'names': ('headers', 'current'),
                'skiprows': 1,
                'index_col': 0,
            }
        }
        return {**base, **special.get(sheet, {})}


def clean_string(
    text: str,
    fill: str = ' ',
    pattern: str = r'\W'
) -> str:
    return fill.join(part for part in re.split(pattern, text.strip()) if part)


def assign_index_by_row_count(df: pd.DataFrame) -> pd.DataFrame:

    index_map = INDEX_MAP

    df.index = index_map.get(df.shape[0], DEFAULT_INDEX)

    return df


def clean_summary_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    rows_to_clean = [
        # Trim Entity
        'insured',
        'address',
        'counterparty',
        'beneficiary_address',
    ] + [
        'loadport_locality',
        'loadport_country',
        'disport_locality',
        'disport_country',
        'vessel'
    ]

    for row in rows_to_clean:
        val = df.at[row, 'current']
        cleaned = clean_string(val).title()
        df.at[row, 'current'] = cleaned

    for row in ['insured', 'counterparty']:
        val = df.at[row, 'current']
        df.at[row, 'current'] = regex_trim_entity(val)

    # Validating Locations
    for row in ['disport_locality', 'disport_country']:
        val = df.at[row, 'current']
        df.at[row, 'current'] = regex_trim_for_orders(val)

    val = df.at['disport_locality', 'current']
    df.at['disport_locality', 'current'] = regex_trim_country(val)

    # Validating Discharge Port Location
    loc = df.at['disport_locality', 'current'].strip().lower()
    cty = df.at['disport_country', 'current'].strip().lower()
    if loc == cty:
        df.at['disport_locality', 'current'] = 'Unknown'

    # Remove Specific Values
    values_to_remove = [
        'Not Disclosed;Not Disclosed',
        'Tba;Tba',
        'Unknown;Unknown',
        'TBA',
        'Tba',
        'Tbn'
    ]

    # Patch Loadports
    TO_REPLACE = {
        'De Kastri': 'De-Kastri',
        'Rostov On Don': 'Rostov-On-Don',
        'Ust Luga': 'Ust-Luga'
    }

    df['current'] = df['current'].replace(values_to_remove, 'Unknown')
    df['current'] = df['current'].replace(TO_REPLACE)

    return df


def standardize_dataset(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardizes the dataset by cleaning column names and dropping empty
    rows and columns.
    """
    return df.pipe(standardize_column_names).pipe(drop_empty_rows_and_columns)


def standardize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardizes the column names by applying a series of transformations:
    1. Trimming whitespaces with a custom trim function.
    2. Transliteration of non-ASCII characters.
    3. Converting the names to lowercase.
    """
    def clean_column_name(col: str) -> str:
        return str.lower(transliterate_to_latin(clean_string(col, fill='_')))

    df.columns = [clean_column_name(col) for col in df.columns]
    return df


def transliterate_to_latin(
    word: str, mapping: dict[str, str] = CYRILLIC_TO_LATIN
) -> str:
    return ''.join(mapping.get(char.lower(), char) for char in word)


def drop_empty_rows_and_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Drops rows and columns that are completely empty.
    """
    return df.dropna(axis=0, how='all').dropna(axis=1, how='all')


def regex_trim(value: str, pattern: str, group_name: str) -> Optional[str]:
    """Generic function to apply regex and extract the named group."""
    if isinstance(value, str):
        match = re.search(pattern, value)
        if match:
            return match.group(group_name)
    return value


def regex_trim_country(value: str) -> Optional[str]:
    pattern = (
        r'One Or \w+\s+Safe\s+Port(?:\s+S|s)?\s+'
        r'(?P<country>.+)$'
    )
    return regex_trim(value, pattern, 'country')


def regex_trim_entity(value: str) -> Optional[str]:
    return regex_trim(value, r'To The Order Of (?P<entity>.*)', 'entity')


def regex_trim_for_orders(value: str) -> Optional[str]:
    return regex_trim(value, r'(?P<country>.*) For Orders', 'country')


class DateService:
    @staticmethod
    def get_date(days_shift: int) -> datetime.date:
        return datetime.date.today() + datetime.timedelta(days=days_shift)

    @staticmethod
    def resolve_eta(eta: datetime.date | None = None, fallback_days: int = 40) -> datetime.date:
        return eta or DateService.get_date(fallback_days)
