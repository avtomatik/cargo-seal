import json
import re

import pandas as pd
from django.conf import settings
from openpyxl import load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from api.serializers import (CoverageSerializer, DocumentSerializer,
                             FormMergeSerializer, PolicySerializer,
                             VesselSerializer)
from coverage.models import Coverage, Policy
from logistics.models import Shipment
from vessels.models import Document, Vessel


class CoverageViewSet(viewsets.ModelViewSet):

    queryset = Coverage.objects.all()
    serializer_class = CoverageSerializer
    parser_classes = [FormParser, MultiPartParser]

    @action(methods=['get', 'post'], detail=True)
    def draft(self, request):
        if request.method == 'GET':
            # TODO: Implement: Produce Coverage Documents
            return Response(
                {'message': 'This Endpoint Functionality Is Under Construction'},
                status=status.HTTP_200_OK
            )
        if request.method == 'POST':
            # TODO: Implement: Translate Form to Documents
            return Response(
                {'message': 'This Endpoint Functionality Is Under Construction'},
                status=status.HTTP_200_OK
            )

    @action(methods=['post'], detail=False)
    def push(self, request):
        """Push Declaration to Database Handler."""
        # =====================================================================
        # curl -X POST -F "file=@path/to/file;type=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" http://localhost:8000/api/coverage/push/
        # =====================================================================
        # =====================================================================
        # TODO: Make It Clear
        # =====================================================================
        file = request.data.get('file')

        if file:
            SHEET_NAMES_EXPECTED = ('declaration_form', 'bl_breakdown')

            sheet_names, version, last_modified_by = self.extract_workbook_data(
                file
            )

            df_frm = pd.read_excel(
                file,
                sheet_name=SHEET_NAMES_EXPECTED[0],
                names=('headers', 'current'),
                index_col=0,
                skiprows=1,
                skipfooter=2,
            ).transpose()

            df_frm.columns = map(
                lambda _: self.trim_string(_, '_').lower(), df_frm.columns
            )

            df_bls = pd.read_excel(
                file,
                sheet_name=SHEET_NAMES_EXPECTED[-1]
            ).dropna(axis=0)

            df_bls.columns = map(
                lambda _: self.trim_string(_, '_').lower(), df_bls.columns
            )

            df_bls['subject_matter_insured'] = df_bls['subject_matter_insured'].apply(
                self.trim_string
            ).apply(str.title)

            df_bls = df_bls.sort_values(by='bl_date')

            df_bls = df_bls.groupby('subject_matter_insured').agg({
                'bl_number': 'count',
                'bl_date': 'max',
                'weight_mt_in_vacuum': 'sum',
                'volume_bbl': 'sum',
                'sum_insured_100_usd': 'sum',
            })

            # df_bls['deal_number'] = deal_number

            # df_bls = df_bls.reset_index().set_index('deal_number')

            print(df_bls)

            with open(settings.BASE_DIR.joinpath('data').joinpath('columns.json')) as file:
                COLUMNS = json.load(file)

            if all(x == y for x, y in zip(df_frm.columns, COLUMNS[version]['expected'])):
                df_frm.columns = COLUMNS[version]['fitted']

            data_received = df_frm.loc['current'].to_dict()
            data_received['operator'] = last_modified_by

            data_received.pop('basis_of_valuation', None)
            data_received.pop('subject_matter_insured', None)
            data_received.pop('_', None)

        # =====================================================================
        # TODO: Validation
        # =====================================================================
            data = {}

            for key, value in data_received.items():
                value_distillated = self.distillate_value(value)
                if value_distillated:
                    data[key] = value_distillated

            return Response(data, status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)

    def distillate_value(self, value):
        if isinstance(value, str):
            string = self.trim_string(value).title()
            if string in ['Not Disclosed', 'Tba', 'Unknown']:
                return
            return string
        return value

    def extract_workbook_data(self, file):
        wb = load_workbook(file, read_only=True, keep_links=False)
        sheet_names = wb.sheetnames
        version = wb['declaration_form']['A1'].value
        last_modified_by = wb.properties.lastModifiedBy
        wb.close()
        return sheet_names, version, last_modified_by

    def trim_string(self, string: str, fill: str = ' ', char: str = r'\W') -> str:
        return fill.join(filter(bool, re.split(char, string)))


class FormMergeViewSet(viewsets.ReadOnlyModelViewSet):

    queryset = Shipment.objects.all()
    serializer_class = FormMergeSerializer


class DocumentViewSet(viewsets.ReadOnlyModelViewSet):

    queryset = Document.objects.all()
    serializer_class = DocumentSerializer


class PolicyViewSet(viewsets.ReadOnlyModelViewSet):

    queryset = Policy.objects.all()
    serializer_class = PolicySerializer


class VesselViewSet(viewsets.ReadOnlyModelViewSet):

    queryset = Vessel.objects.all()
    serializer_class = VesselSerializer

    @action(methods=['post'], detail=True)
    def check(self, request):
        # TODO: Implement
        return Response(
            {'message': 'This Endpoint Functionality Is Under Construction'},
            status=status.HTTP_200_OK
        )
